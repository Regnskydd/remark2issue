#!/usr/bin/env python3

def die(message):
    print(message)
    import sys
    sys.exit(1)

from remark import Remark
import optparse
import json

import getpass

try:
    import requests
    from requests.auth import HTTPBasicAuth
except ImportError:
    die("Please install restkit")

try:
    from openpyxl import load_workbook
except ImportError:
    die("Please install openpyxl")

# CONSTANTS
IR_SHEET = 2
ID_COL = 1
AUTHOR_COL = 5
COMMENT_COL = 6
DECISION_COL = 9
DECISION_COMMENT_COL = 10
ACTION_DESCRIPTION_COL = 11
STATUS_COL = 12
#BEARER_TOKEN = "ADD TOKEN"

def parse_args():
    parser = optparse.OptionParser()
    parser.add_option('-u', '--user', dest='user', default=getpass.getuser(), help='Username to access JIRA')
    parser.add_option('-p', '--password', dest='password', help='Password to access JIRA')
    parser.add_option('-j', '--jira', dest='jira_url', default='http://localhost:8080', help='JIRA Base URL')
    parser.add_option('-f', '--file', dest='filename', help='Filename for excel workbook')
    parser.add_option('-k', '--key', dest='key', help='Project key to jira')
    parser.add_option('-s', '--status', dest='status', help='Update status on actions based on issues in JIRA')
    parser.add_option('-i', '--issue', dest='issue', help='Create issues based on remarks in excel sheet')

    return parser.parse_args()

def get_password():
    return getpass.getpass("Please enter the Jira Password:")

def fetch_open_remarks(workbook):
    sheet = workbook[workbook.sheetnames[IR_SHEET]]
    remarks = []
    
    for row_index in range(sheet.max_row)[1:]:
        # Check if the remark have been accepted or postponed
        if sheet.cell(row_index,DECISION_COL).value == 'A' or sheet.cell(row_index,DECISION_COL).value == 'P':
            # Check if the remark is open
            if not sheet.cell(row_index,STATUS_COL).value:
                # Check if we already created an issue in order to avoid creating duplicates
                if not sheet.cell(row_index,ACTION_DESCRIPTION_COL).value:
                    remarks.append(Remark(row_index,
                                          sheet.cell(row_index,AUTHOR_COL).value,
                                          sheet.cell(row_index,COMMENT_COL).value,
                                          sheet.cell(row_index,DECISION_COL).value,
                                          sheet.cell(row_index,DECISION_COMMENT_COL).value,
                                          sheet.cell(row_index,ACTION_DESCRIPTION_COL).value,
                                          sheet.cell(row_index,STATUS_COL).value))
    return remarks

# Fetches the status for an issue
def fetch_issue_status_from_jira(options,auth):
    #headers = {"Content-Type": "application/json", "Authorization": BEARER_TOKEN}
    headers = {"Content-Type": "application/json"}
    
    workbook = load_workbook(options.filename)
    sheet = workbook[workbook.sheetnames[IR_SHEET]]
    for row_index in range(sheet.max_row)[1:]:
        # Check if the remark have been accepted or postponed
        if sheet.cell(row_index,DECISION_COL).value == 'A' or sheet.cell(row_index,DECISION_COL).value == 'P':
            #Check if the remark is open
            if not sheet.cell(row_index,STATUS_COL).value:
                action_description = sheet.cell(row_index,ACTION_DESCRIPTION_COL).value
                # This could be empty if no issues have been created with the script previously or if the user have manually removed them. So double check to be certain. 
                if action_description != None:
                    # Get the JIRA key we wrote in the cell when the issue was created
                    request_url = options.jira_url + "/rest/api/latest/issue/" + action_description.split(' ', 1)[0]
                    #response = requests.get(url=request_url,headers=headers,data=json.dumps(payload))
                    response = requests.get(url=request_url,auth=auth,headers=headers)
                    issue = json.loads(response.content)
                    
                    if response.status_code == 200 or response.status_code == 201:
                        # We only care to update issues that are closed in JIRA
                        if issue['fields']['status']['name'] == 'Done':
                            comments = issue['fields']['comment']['comments']
                            if comments:
                                # Should add support for multiple comments
                                sheet.cell(row_index,ACTION_DESCRIPTION_COL).value = comments[0]['author']['name'] + ': ' + comments[0]['body']
                                sheet.cell(row_index,STATUS_COL).value = 'F'
                                #print('Update action description on row', row_index, ' from issue ', json.dumps(issue['key']))
                                print('Update action description on row ' + str(row_index) + ' from issue ' + json.dumps(issue['key']))
                            else:
                                print('You should add comments to your Jira issue when closing it, now you have to manually update the excel or add a comment in jira and re-run this script')
                    else:
                        die(response)
    workbook.save(options.filename)                  
# Writes the JIRA issue key that was created for an issue                        
def write_issue_key_on_remark(remark,options,response):
    workbook = load_workbook(options.filename)
    sheet = workbook[workbook.sheetnames[IR_SHEET]]
    
    sheet.cell(remark.get_identifier(),ACTION_DESCRIPTION_COL).value = json.loads(response.content)['key'] + " have been created, but not yet completed. This cell will be updated when task is completed."
    workbook.save(options.filename)
    print('Created JIRA issue: ' + json.loads(response.content)['key'])

def create_issue(remark, options,auth):
    request_url = options.jira_url + "/rest/api/latest/issue/"
    payload = {"fields": {"project": {"key": options.key},"summary": "IR-Remark","description": remark.get_comment(), "issuetype": {"name": "Task"}}}
    #headers = {"Content-Type": "application/json", "Authorization": BEARER_TOKEN}
    #response = requests.post(url=request_url,headers=headers,data=json.dumps(payload))
    headers = {"Content-Type": "application/json"}
    response = requests.post(url=request_url,auth=auth,headers=headers,data=json.dumps(payload))

    if response.status_code == 200 or response.status_code == 201:
        write_issue_key_on_remark(remark,options,response)
    else:
        die(response.content)
    
if __name__ == '__main__':
    (options, args) = parse_args()
    
    # Exit if no file was supplied to program
    if not options.filename:
        parser.error('Filename to workbook not given')

    # Exit if no project key was supplied to program
    if not options.key:
        parser.error('Project key to JIRA not given')
    
    if options.issue:
        # Basic Auth is usually easier for scripts like this to deal with than Cookies.
        auth = HTTPBasicAuth(options.user, options.password or get_password())

        workbook = load_workbook(options.filename)
    
        remarks = fetch_open_remarks(workbook)

        for remark in remarks:
            create_issue(remark,options,auth)

    if options.status:
        fetch_issue_status_from_jira(options,auth)
